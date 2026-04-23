"""
Generate the Week 1 sample workbook from the checked-in sample cohort JSON.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
SAMPLE_JSON = ROOT / "app" / "src" / "data" / "cohort_data.sample.json"
OUTPUT_XLSX = ROOT / "etl" / "sample_data" / "sample_cohort.xlsx"


def main() -> None:
    with open(SAMPLE_JSON, "r", encoding="utf-8") as file_handle:
        data = json.load(file_handle)

    workbook = Workbook()
    self_sheet = workbook.active
    self_sheet.title = "Self-Eval"
    peer_sheet = workbook.create_sheet("Peer-Eval")
    staff_sheet = workbook.create_sheet("Staff-Eval")
    gates_sheet = workbook.create_sheet("Gates")

    score_headers = [f"score_{index}" for index in range(1, 23)]

    self_sheet.append(["student_id", "student_name", "team", *score_headers])
    peer_sheet.append(
        ["evaluator_id", "subject_id", *score_headers, "cross_team_flag"]
    )
    staff_sheet.append(["evaluator_id", "subject_id", *score_headers])
    gates_sheet.append(
        ["student_id", "sprints_complete", "demo_day", "negotiation_sim"]
    )

    for student in data["students"]:
        self_sheet.append(
            [
                student["id"],
                student["name"],
                student["team"],
                *student["scores"]["self_scores"],
            ]
        )

        max_peer_count = max(student["scores"]["peer_count"])
        for evaluator_index in range(max_peer_count):
            peer_scores = []
            for score, count in zip(
                student["scores"]["peer_avg"],
                student["scores"]["peer_count"],
            ):
                peer_scores.append(score if evaluator_index < count else None)
            peer_sheet.append(
                [
                    f"peer_{evaluator_index + 1}",
                    student["id"],
                    *peer_scores,
                    False,
                ]
            )

        max_staff_count = max(student["scores"]["staff_count"])
        for evaluator_index in range(max_staff_count):
            staff_scores = []
            for score, count in zip(
                student["scores"]["staff_avg"],
                student["scores"]["staff_count"],
            ):
                staff_scores.append(score if evaluator_index < count else None)
            staff_sheet.append(
                [f"staff_{evaluator_index + 1}", student["id"], *staff_scores]
            )

        gates_sheet.append(
            [
                student["id"],
                student["certification"]["sprints_complete"],
                student["certification"]["demo_day"],
                student["certification"]["negotiation_sim"],
            ]
        )

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    print(f"Wrote sample workbook: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
