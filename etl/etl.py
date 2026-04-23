"""
FLI Scorecard ETL
==================
Processes Google Form evaluation exports (.xlsx) into cohort_data.json.

Usage:
    python etl.py path/to/evaluations.xlsx

Output:
    cohort_data.json -> copy to app/src/data/cohort_data.json

Sheet naming convention expected in .xlsx:
    "Self-Eval"  - self-evaluation responses
    "Peer-Eval"  - peer evaluation responses
    "Staff-Eval" - staff evaluation responses
    "Gates"      - sprint completion, demo day, negotiation sim booleans
"""

import json
import sys
from collections import defaultdict
from dataclasses import asdict
from pathlib import Path

import openpyxl

from schema import (
    CERTIFICATION_SCORE_THRESHOLD,
    CLUSTER_BOUNDARIES,
    CLUSTER_KEYS,
    CLUSTER_NAMES,
    CLUSTER_WEIGHTS,
    COMPETENCIES,
    PERCEPTION_GAP_THRESHOLD,
    STAFF_DIVERGENCE_THRESHOLD,
    CertificationRecord,
    CertificationStatus,
    ClusterAverages,
    ClusterRanking,
    CohortAnalytics,
    CohortData,
    CohortSummary,
    PerceptionGapFlag,
    StaffDivergenceFlag,
    Student,
    StudentFlags,
    StudentScores,
)


SHEET_SELF_EVAL = "Self-Eval"
SHEET_PEER_EVAL = "Peer-Eval"
SHEET_STAFF_EVAL = "Staff-Eval"
SHEET_GATES = "Gates"

SCORE_COUNT = 22


def main(xlsx_path: str) -> None:
    path = Path(xlsx_path)
    if not path.exists():
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Processing: {path.resolve()}")
    wb = openpyxl.load_workbook(path, data_only=True)

    students = parse_evaluations(wb)
    cohort_analytics = compute_cohort_analytics(students)

    output = CohortData(
        cohort=CohortSummary(
            name="Spring 2026 - Brownsville",
            students=len(students),
            certified=sum(
                1
                for student in students
                if student.certification.status == CertificationStatus.CERTIFIED
            ),
            avg_weighted_score=round(
                sum(student.weighted_overall for student in students) / len(students),
                2,
            ),
        ),
        competencies=COMPETENCIES,
        students=students,
        cohort_analytics=cohort_analytics,
    )

    out_path = Path("cohort_data.json")
    with open(out_path, "w", encoding="utf-8") as file_handle:
        json.dump(asdict(output), file_handle, indent=2)

    print(f"\nOutput written to: {out_path}")
    print(f"  Students processed : {output.cohort.students}")
    print(f"  Certified          : {output.cohort.certified}")
    print(f"  Avg weighted score : {output.cohort.avg_weighted_score}")
    print("\nNext step: copy cohort_data.json -> app/src/data/cohort_data.json")


def parse_evaluations(wb: openpyxl.Workbook) -> list[Student]:
    """
    Parse all evaluation sheets and return a list of fully computed Students.

    Expected sheet structure:
    - Self-Eval:  [student_id, student_name, team, score_1...score_22]
    - Peer-Eval:  [evaluator_id, subject_id, score_1...score_22, cross_team_flag]
    - Staff-Eval: [evaluator_id, subject_id, score_1...score_22]
    - Gates:      [student_id, sprints_complete, demo_day, negotiation_sim]
    """
    for sheet_name in (
        SHEET_SELF_EVAL,
        SHEET_PEER_EVAL,
        SHEET_STAFF_EVAL,
        SHEET_GATES,
    ):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing required sheet: {sheet_name}")

    roster: dict[str, dict[str, str | list[float]]] = {}
    peer_rows_by_student: dict[str, list[list[float | None]]] = defaultdict(list)
    staff_rows_by_student: dict[str, list[list[float | None]]] = defaultdict(list)
    gates_by_student: dict[str, tuple[bool, bool, bool]] = {}

    for row in wb[SHEET_SELF_EVAL].iter_rows(min_row=2, values_only=True):
        student_id = safe_str(row[0]) if len(row) > 0 else ""
        if not student_id:
            continue

        roster[student_id] = {
            "name": safe_str(row[1]) if len(row) > 1 else student_id,
            "team": safe_str(row[2]) if len(row) > 2 else "Unassigned",
            "self_scores": coalesce_score_row(
                normalize_score_row(row[3 : 3 + SCORE_COUNT])
            ),
        }

    for row in wb[SHEET_PEER_EVAL].iter_rows(min_row=2, values_only=True):
        subject_id = safe_str(row[1]) if len(row) > 1 else ""
        if not subject_id:
            continue
        peer_rows_by_student[subject_id].append(
            normalize_score_row(row[2 : 2 + SCORE_COUNT])
        )

    for row in wb[SHEET_STAFF_EVAL].iter_rows(min_row=2, values_only=True):
        subject_id = safe_str(row[1]) if len(row) > 1 else ""
        if not subject_id:
            continue
        staff_rows_by_student[subject_id].append(
            normalize_score_row(row[2 : 2 + SCORE_COUNT])
        )

    for row in wb[SHEET_GATES].iter_rows(min_row=2, values_only=True):
        student_id = safe_str(row[0]) if len(row) > 0 else ""
        if not student_id:
            continue

        gates_by_student[student_id] = (
            parse_bool(row[1] if len(row) > 1 else None),
            parse_bool(row[2] if len(row) > 2 else None),
            parse_bool(row[3] if len(row) > 3 else None),
        )

    students: list[Student] = []
    for student_id in sorted(roster.keys()):
        student_meta = roster[student_id]
        self_scores = student_meta["self_scores"]
        if not isinstance(self_scores, list):
            raise TypeError(f"Invalid self score payload for {student_id}")

        peer_score_rows = peer_rows_by_student.get(student_id, [])
        staff_score_rows = staff_rows_by_student.get(student_id, [])
        peer_avg, peer_count = aggregate_scores(peer_score_rows)
        staff_avg, staff_count = aggregate_scores(staff_score_rows)
        cluster_averages = compute_cluster_averages(peer_avg)
        weighted_overall = compute_weighted_overall(cluster_averages)
        sprints_complete, demo_day, negotiation_sim = gates_by_student.get(
            student_id,
            (False, False, False),
        )

        students.append(
            Student(
                id=student_id,
                name=str(student_meta["name"]),
                team=str(student_meta["team"]),
                scores=StudentScores(
                    self_scores=self_scores,
                    peer_avg=peer_avg,
                    peer_count=peer_count,
                    staff_avg=staff_avg,
                    staff_count=staff_count,
                ),
                cluster_averages=cluster_averages,
                weighted_overall=weighted_overall,
                certification=compute_certification(
                    weighted_overall,
                    sprints_complete,
                    demo_day,
                    negotiation_sim,
                ),
                flags=StudentFlags(
                    perception_gaps=compute_perception_gaps(self_scores, peer_avg),
                    staff_divergence=compute_staff_divergence(staff_score_rows),
                ),
            )
        )

    return students


def safe_str(value) -> str:
    """Convert workbook cell values to trimmed strings; blank/null becomes empty."""
    if value is None:
        return ""
    return str(value).strip()


def parse_bool(value) -> bool:
    """Parse spreadsheet booleans and common truthy/falsey strings safely."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False

    normalized = str(value).strip().lower()
    if normalized in {"true", "t", "yes", "y", "1", "complete", "completed"}:
        return True
    if normalized in {"false", "f", "no", "n", "0", "", "incomplete"}:
        return False
    return False


def safe_float(value) -> float | None:
    """Safely convert a cell value to float; return None if blank or invalid."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def normalize_score_row(values) -> list[float | None]:
    """Normalize a partial row of competency scores into a length-22 list."""
    scores = [safe_float(value) for value in list(values)[:SCORE_COUNT]]
    if len(scores) < SCORE_COUNT:
        scores.extend([None] * (SCORE_COUNT - len(scores)))
    return scores


def coalesce_score_row(values: list[float | None]) -> list[float]:
    """Convert optional score rows into numeric rows for schema fields that require floats."""
    return [value if value is not None else 0.0 for value in values]


def aggregate_scores(
    score_rows: list[list[float | None]],
) -> tuple[list[float], list[int]]:
    """
    Given a list of evaluator rows (each row is 22 scores, some may be None),
    return (avg_per_competency[22], count_per_competency[22]).
    """
    avgs = []
    counts = []
    for index in range(SCORE_COUNT):
        col_scores = [row[index] for row in score_rows if row[index] is not None]
        if col_scores:
            avgs.append(round(sum(col_scores) / len(col_scores), 3))
            counts.append(len(col_scores))
        else:
            avgs.append(0.0)
            counts.append(0)
    return avgs, counts


def compute_cluster_averages(scores: list[float]) -> ClusterAverages:
    """Compute per-cluster averages from a length-22 score array."""
    avgs = [
        round(sum(scores[start:end]) / (end - start), 3)
        for start, end in CLUSTER_BOUNDARIES
    ]
    return ClusterAverages(
        ai_digital=avgs[0],
        product=avgs[1],
        leadership=avgs[2],
        professional=avgs[3],
        community=avgs[4],
    )


def compute_weighted_overall(cluster_avgs: ClusterAverages) -> float:
    """Weighted average across 5 clusters using locked cluster weights."""
    values = [
        cluster_avgs.ai_digital,
        cluster_avgs.product,
        cluster_avgs.leadership,
        cluster_avgs.professional,
        cluster_avgs.community,
    ]
    return round(sum(value * weight for value, weight in zip(values, CLUSTER_WEIGHTS)), 3)


def compute_certification(
    weighted_overall: float,
    sprints_complete: bool,
    demo_day: bool,
    negotiation_sim: bool,
) -> CertificationRecord:
    """Apply certification business rules."""
    meets_score_bar = weighted_overall >= CERTIFICATION_SCORE_THRESHOLD
    all_gates = sprints_complete and demo_day and negotiation_sim

    if meets_score_bar and all_gates:
        status = CertificationStatus.CERTIFIED
    elif meets_score_bar and not all_gates:
        status = CertificationStatus.CONDITIONAL
    else:
        status = CertificationStatus.NOT_CERTIFIED

    return CertificationRecord(
        status=status,
        meets_score_bar=meets_score_bar,
        sprints_complete=sprints_complete,
        demo_day=demo_day,
        negotiation_sim=negotiation_sim,
    )


def compute_perception_gaps(
    self_scores: list[float],
    peer_avg: list[float],
) -> list[PerceptionGapFlag]:
    """Flag competencies where self-score exceeds peer avg by >= threshold."""
    gaps = []
    for competency in COMPETENCIES:
        index = competency.id - 1
        gap = self_scores[index] - peer_avg[index]
        if gap >= PERCEPTION_GAP_THRESHOLD:
            gaps.append(
                PerceptionGapFlag(
                    competency_id=competency.id,
                    competency_name=competency.name,
                    self_score=self_scores[index],
                    peer_avg=peer_avg[index],
                    gap=round(gap, 2),
                )
            )
    return gaps


def compute_staff_divergence(
    staff_score_rows: list[list[float | None]],
) -> list[StaffDivergenceFlag]:
    """Flag competencies where staff raters diverge by >= threshold."""
    divergences = []
    for competency in COMPETENCIES:
        index = competency.id - 1
        scores = [row[index] for row in staff_score_rows if row[index] is not None]
        if len(scores) < 2:
            continue

        divergence = max(scores) - min(scores)
        if divergence >= STAFF_DIVERGENCE_THRESHOLD:
            divergences.append(
                StaffDivergenceFlag(
                    competency_id=competency.id,
                    competency_name=competency.name,
                    staff_scores=scores,
                    divergence=round(divergence, 2),
                )
            )
    return divergences


def compute_cohort_analytics(students: list[Student]) -> CohortAnalytics:
    """Compute program-wide analytics from student peer averages and cluster scores."""
    if not students:
        return CohortAnalytics(
            strongest_competencies=[],
            weakest_competencies=[],
            avg_self_vs_peer_gap=0.0,
            cluster_rankings=[],
        )

    competency_peer_avgs: list[tuple[str, float]] = []
    for competency in COMPETENCIES:
        index = competency.id - 1
        avg_score = round(
            sum(student.scores.peer_avg[index] for student in students) / len(students),
            3,
        )
        competency_peer_avgs.append((competency.name, avg_score))

    strongest_competencies = [
        name
        for name, _ in sorted(
            competency_peer_avgs,
            key=lambda item: (-item[1], item[0]),
        )[:3]
    ]
    weakest_competencies = [
        name
        for name, _ in sorted(
            competency_peer_avgs,
            key=lambda item: (item[1], item[0]),
        )[:3]
    ]

    avg_self_vs_peer_gap = round(
        sum(
            self_score - peer_score
            for student in students
            for self_score, peer_score in zip(
                student.scores.self_scores,
                student.scores.peer_avg,
            )
        ) / (len(students) * SCORE_COUNT),
        2,
    )

    cluster_scores = [
        (
            cluster_name,
            round(
                sum(getattr(student.cluster_averages, cluster_key) for student in students)
                / len(students),
                2,
            ),
        )
        for cluster_name, cluster_key in zip(CLUSTER_NAMES, CLUSTER_KEYS)
    ]

    cluster_rankings = [
        ClusterRanking(cluster=cluster, avg_score=avg_score, rank=index + 1)
        for index, (cluster, avg_score) in enumerate(
            sorted(cluster_scores, key=lambda item: (-item[1], item[0]))
        )
    ]

    return CohortAnalytics(
        strongest_competencies=strongest_competencies,
        weakest_competencies=weakest_competencies,
        avg_self_vs_peer_gap=avg_self_vs_peer_gap,
        cluster_rankings=cluster_rankings,
    )


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python etl.py path/to/evaluations.xlsx")
        sys.exit(1)
    main(sys.argv[1])
